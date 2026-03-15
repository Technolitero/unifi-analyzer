"""
UniFi Analyzer — FastAPI Backend
Run with:
  source .venv/bin/activate  # Activate virtual environment
  uvicorn main:app --host 0.0.0.0 --port 8080 --reload
Then open http://localhost:8080 in your browser.
"""

import asyncio
import io
import json
import re
import tempfile
import traceback
import uuid
import zipfile
from pathlib import Path
from typing import Dict, List, Optional

import queue
import threading

from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from unifi_client import UnifiClient
from config_analyzer import ConfigAnalyzer
from network_optimizer import NetworkOptimizer
from pcap_handler import PcapCapture, format_pcap_for_ai
from credentials import load_config, save_config
from config_export import run_export

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = FastAPI(title="UniFi Analyzer", version="1.0.0")

STATIC_DIR = Path(__file__).parent / "static"
HISTORY_DIR = Path(__file__).parent / "history"

# In-memory store for raw PCAP bytes — avoids base64 OOM in browser
_pcap_cache: Dict[str, bytes] = {}

# In-memory store for completed config export jobs: job_id -> {"out_dir": Path, "exit_code": int}
_export_jobs: Dict[str, dict] = {}


@app.get("/", response_class=HTMLResponse)
async def root():
    html_path = STATIC_DIR / "index.html"
    if html_path.exists():
        return HTMLResponse(content=html_path.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="Frontend not found")


# ---------------------------------------------------------------------------
# Request models
# ---------------------------------------------------------------------------

class ConnectionParams(BaseModel):
    host: str
    username: str
    password: str
    port: int = 443
    site: str = "default"


class PcapParams(BaseModel):
    host: str
    ssh_username: str
    ssh_password: Optional[str] = None
    ssh_key_path: Optional[str] = None
    ssh_port: int = 22
    interface: str = "eth4"
    duration: int = 30
    packet_count: Optional[int] = None
    bpf_filter: str = ""
    description: str = ""
    max_display_packets: int = 200


class PcapZipEntry(BaseModel):
    download_id: str
    filename: str  # e.g. "RouterA_eth4_2026-03-07T14-30-00.pcap"


class PcapZipRequest(BaseModel):
    files: List[PcapZipEntry]
    zip_name: str = "captures.zip"


class InterfaceFetchParams(BaseModel):
    host: str
    ssh_username: str
    ssh_password: Optional[str] = None
    ssh_key_path: Optional[str] = None
    ssh_port: int = 22
    # Optional UniFi API creds to enrich VLAN subinterfaces with network names
    api_username: Optional[str] = None
    api_password: Optional[str] = None
    api_port: int = 443
    api_site: str = "default"


class DeviceListParams(BaseModel):
    host: str
    username: str
    password: str
    port: int = 443
    site: str = "default"


class AllInterfacesParams(BaseModel):
    udm_host: str
    api_username: str
    api_password: str
    api_port: int = 443
    api_site: str = "default"
    # UDM/gateway SSH — username is always root on UniFiOS
    udm_ssh_username: str = "root"
    udm_ssh_password: Optional[str] = None
    udm_ssh_key_path: Optional[str] = None
    # Switch / AP SSH
    device_ssh_username: str = "admin"
    device_ssh_password: Optional[str] = None
    device_ssh_key_path: Optional[str] = None
    ssh_port: int = 22


class SavedConfig(BaseModel):
    host: str = ""
    api_port: int = 443
    api_site: str = "default"
    api_username: str = ""
    api_password: str = ""

    ssh_port: int = 22
    udm_ssh_username: str = "root"
    udm_ssh_password: str = ""
    device_ssh_username: str = "admin"
    device_ssh_password: str = ""


# ---------------------------------------------------------------------------
# API Routes
# ---------------------------------------------------------------------------

@app.post("/api/analyze")
async def analyze(params: ConnectionParams):
    """Connect to UDM, retrieve config, and return analysis suggestions."""
    client = UnifiClient(
        host=params.host,
        username=params.username,
        password=params.password,
        port=params.port,
        site=params.site,
    )
    try:
        client.login()
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Authentication failed: {e}")

    try:
        config = client.get_all_config()
        # Enrich with zone-based firewall data (best-effort, newer UniFiOS only)
        try:
            config["firewall_zones"] = client.get_firewall_zones_v2_session()
        except Exception:
            config["firewall_zones"] = []
        try:
            config["firewall_policies"] = client.get_firewall_policies_session()
        except Exception:
            config["firewall_policies"] = []
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to retrieve config: {e}")
    finally:
        client.logout()

    # Run both analyzers and merge results
    analyzer = ConfigAnalyzer(config)
    base_suggestions = analyzer.analyze()

    optimizer = NetworkOptimizer(config)
    opt_result = optimizer.run()

    # Map optimizer 3-tier severity to the UI's 5-tier vocabulary
    _sev_map = {"critical": "critical", "recommended": "medium", "informational": "info"}

    # Deduplicate: normalize titles by stripping quoted device/SSID names
    def _norm(title: str) -> str:
        return re.sub(r"'[^']*'", "*", title.lower()).strip()

    seen_titles = {_norm(s["title"]) for s in base_suggestions}
    opt_suggestions = []
    for issue in opt_result["issues"]:
        mapped_sev = _sev_map.get(issue["severity"], "info")
        norm = _norm(issue["title"])
        if norm in seen_titles:
            continue
        seen_titles.add(norm)
        opt_suggestions.append({
            "category":       issue["category"],
            "severity":       mapped_sev,
            "title":          issue["title"],
            "detail":         issue["detail"],
            "recommendation": issue["recommendation"],
        })

    _sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
    suggestions = sorted(
        base_suggestions + opt_suggestions,
        key=lambda s: _sev_order.get(s["severity"], 9),
    )

    security_score = {
        "score":       opt_result["score"],
        "label":       opt_result["score_label"],
        "posture":     opt_result["posture"],
        "description": opt_result["score_description"],
        "counts":      opt_result["issue_counts"],
        "hardening":   opt_result["hardening_measures"],
    }

    def _fmt_dev_type(raw_type: str) -> str:
        t = (raw_type or "").lower()
        if t.startswith("usw") or t.startswith("ubb"): return "Switch"
        if t.startswith("uap") or t.startswith("uwa"): return "AP"
        if t.startswith("udm") or t.startswith("uxg") or t.startswith("usg") or t in ("ugw",): return "Gateway"
        return raw_type or "Unknown"

    def _summarize_devices(raw: list) -> list:
        out = []
        for d in raw:
            raw_type = (d.get("type") or "").lower()
            is_ap = raw_type.startswith("uap") or raw_type.startswith("uwa")
            # Mesh enabled: AP has mesh VAP active or a wireless uplink
            mesh_enabled = False
            if is_ap:
                mesh_enabled = bool(d.get("mesh_sta_vap_enabled")) or \
                               (d.get("uplink") or {}).get("type") == "wireless"
            out.append({
                "name": d.get("name") or d.get("hostname", ""),
                "ip": d.get("ip", ""),
                "mac": d.get("mac", ""),
                "type": _fmt_dev_type(raw_type),
                "model": d.get("model", ""),
                "version": d.get("version", ""),
                "connected": d.get("state", 0) == 1,
                "mesh_enabled": mesh_enabled,
            })
        return out

    def _summarize_clients(raw: list) -> list:
        out = []
        for c in raw:
            out.append({
                "hostname": c.get("hostname") or c.get("name", ""),
                "ip": c.get("ip", ""),
                "mac": c.get("mac", ""),
                "wired": bool(c.get("is_wired")),
                "network": c.get("network", ""),
                "signal": c.get("signal"),
                "tx_bytes": c.get("tx_bytes", 0),
                "rx_bytes": c.get("rx_bytes", 0),
            })
        return out

    def _summarize_networks(raw: list) -> list:
        out = []
        for n in raw:
            name = n.get("name", "")
            if not name:
                continue
            out.append({
                "name": name,
                "purpose": n.get("purpose", ""),
                "subnet": n.get("ip_subnet", ""),
                "vlan": n.get("vlan"),
                "enabled": n.get("enabled", True),
            })
        return out

    def _summarize_wlans(raw: list) -> list:
        out = []
        for w in raw:
            out.append({
                "ssid": w.get("name", ""),
                "security": w.get("security", ""),
                "enabled": w.get("enabled", True),
                "band": w.get("wlan_band", ""),
                "hidden": bool(w.get("hide_ssid")),
            })
        return out

    filtered_networks = _summarize_networks(config.get("networks", []))

    return JSONResponse({
        "device_count": len(config.get("devices", [])),
        "client_count": len(config.get("clients", [])),
        "network_count": len(filtered_networks),
        "wlan_count": len(config.get("wlans", [])),
        "suggestion_count": len(suggestions),
        "suggestions": suggestions,
        "security_score": security_score,
        "devices": _summarize_devices(config.get("devices", [])),
        "clients": _summarize_clients(config.get("clients", [])),
        "networks": filtered_networks,
        "wlans": _summarize_wlans(config.get("wlans", [])),
    })


@app.post("/api/debug/raw")
async def debug_raw(params: ConnectionParams):
    """Return raw settings and a sample device object for field-name debugging."""
    client = UnifiClient(
        host=params.host, username=params.username,
        password=params.password, port=params.port, site=params.site,
    )
    try:
        client.login()
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Authentication failed: {e}")
    try:
        settings = client.get_settings()
        devices  = client.get_devices()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to retrieve data: {e}")
    finally:
        client.logout()

    return JSONResponse({
        "settings": settings,
        "device_sample": devices[:3],   # first 3 devices only
    })


@app.post("/api/pcap")
async def pcap_capture(params: PcapParams):
    """SSH into UDM, capture packets, and return formatted text for AI."""
    capturer = PcapCapture(
        host=params.host,
        username=params.ssh_username,
        password=params.ssh_password,
        key_path=params.ssh_key_path,
        port=params.ssh_port,
    )

    try:
        loop = asyncio.get_event_loop()
        raw_pcap = await loop.run_in_executor(
            None,
            lambda: capturer.capture(
                interface=params.interface,
                duration=params.duration,
                packet_count=params.packet_count,
                bpf_filter=params.bpf_filter,
            ),
        )
    except Exception as e:
        raise HTTPException(
            status_code=502,
            detail=f"PCAP capture failed: {e}\n{traceback.format_exc()}",
        )

    description = params.description or f"Capture on {params.host} interface {params.interface}"

    text = format_pcap_for_ai(
        raw_pcap=raw_pcap,
        description=description,
        interface=params.interface,
        max_packets=params.max_display_packets,
    )

    full_text = format_pcap_for_ai(
        raw_pcap=raw_pcap,
        description=description,
        interface=params.interface,
        max_packets=0,
    )

    download_id = str(uuid.uuid4())
    _pcap_cache[download_id] = raw_pcap

    return JSONResponse({
        "pcap_size_bytes": len(raw_pcap),
        "download_id": download_id,
        "text": text,
        "full_text": full_text,
    })


@app.get("/api/pcap/download/{download_id}")
async def download_pcap(download_id: str):
    """Stream a previously captured raw PCAP file directly to the browser."""
    raw = _pcap_cache.get(download_id)
    if raw is None:
        raise HTTPException(status_code=404, detail="PCAP not found or expired")

    async def _stream():
        yield raw

    return StreamingResponse(
        _stream(),
        media_type="application/vnd.tcpdump.pcap",
        headers={"Content-Disposition": f'attachment; filename="capture_{download_id[:8]}.pcap"'},
    )


@app.post("/api/pcap/download/zip")
async def download_pcap_zip(req: PcapZipRequest):
    """Bundle multiple stored PCAPs into a zip and stream it to the browser."""
    missing = [f.filename for f in req.files if f.download_id not in _pcap_cache]
    if missing:
        raise HTTPException(status_code=404, detail=f"PCAP(s) not found or expired: {', '.join(missing)}")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for entry in req.files:
            zf.writestr(entry.filename, _pcap_cache[entry.download_id])
    zip_bytes = buf.getvalue()

    async def _stream():
        yield zip_bytes

    safe_name = req.zip_name.replace('"', '')
    return StreamingResponse(
        _stream(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


@app.post("/api/pcap/interfaces")
async def fetch_live_interfaces(params: InterfaceFetchParams):
    """SSH into UDM and return all available network interfaces with metadata."""
    capturer = PcapCapture(
        host=params.host,
        username=params.ssh_username,
        password=params.ssh_password,
        key_path=params.ssh_key_path,
        port=params.ssh_port,
    )

    # Build VLAN ID -> network name map from UniFi API if credentials provided
    vlan_map: dict = {}
    if params.api_username and params.api_password:
        try:
            api_client = UnifiClient(
                host=params.host,
                username=params.api_username,
                password=params.api_password,
                port=params.api_port,
                site=params.api_site,
            )
            api_client.login()
            networks = api_client.get_network_conf()
            api_client.logout()
            for net in networks:
                vid = net.get("vlan")
                name = net.get("name", "")
                if vid and name:
                    vlan_map[int(vid)] = name
        except Exception:
            pass  # API enrichment is best-effort

    try:
        interfaces = capturer.fetch_interfaces(vlan_map=vlan_map or None)
    except Exception as e:
        raise HTTPException(
            status_code=502,
            detail=f"Could not connect via SSH: {e}",
        )

    return JSONResponse({"interfaces": interfaces, "vlan_map": vlan_map})


@app.post("/api/pcap/all-interfaces")
async def fetch_all_interfaces(params: AllInterfacesParams):
    """Discover all managed devices via UniFi API, SSH into each in parallel, return combined interface list."""
    # 1. Get all connected devices from the UniFi controller
    api_client = UnifiClient(
        host=params.udm_host,
        username=params.api_username,
        password=params.api_password,
        port=params.api_port,
        site=params.api_site,
    )
    try:
        api_client.login()
        raw_devices = api_client.get_devices()
        networks = api_client.get_network_conf()
        raw_clients = api_client.get_clients()
        api_client.logout()
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"UniFi API login failed: {e}")

    # Build (sw_mac_lower, sw_port_int) -> list of client summaries
    # sw_mac / sw_port are set on wired clients connected to switches/gateway ports
    _client_by_port: dict[tuple, list] = {}
    for c in raw_clients:
        sw_mac = (c.get("sw_mac") or "").lower().strip()
        sw_port = c.get("sw_port")
        if not sw_mac or sw_port is None:
            continue
        try:
            sw_port = int(sw_port)
        except (ValueError, TypeError):
            continue
        key = (sw_mac, sw_port)
        _client_by_port.setdefault(key, []).append({
            "hostname": c.get("hostname") or c.get("name") or "",
            "ip": c.get("ip", ""),
            "mac": c.get("mac", ""),
            "wired": bool(c.get("is_wired")),
        })

    # Build VLAN ID -> network name map
    vlan_map: dict = {}
    for net in networks:
        vid = net.get("vlan")
        name = net.get("name", "")
        if vid and name:
            vlan_map[int(vid)] = name

    def _classify_device_type(raw_type: str) -> str:
        t = (raw_type or "").lower()
        if t.startswith("usw") or t.startswith("ubb"):
            return "switch"
        if t.startswith("uap") or t.startswith("uwa"):
            return "ap"
        if t in ("ugw", "udm", "udmpro", "uxg", "usg"):
            return "gateway"
        if t.startswith("udm") or t.startswith("uxg") or t.startswith("usg"):
            return "gateway"
        return t or "unknown"

    # Build device_mac_lower -> device summary (for device-to-device uplink info)
    _device_by_mac: dict[str, dict] = {}
    for d in raw_devices:
        dmac = (d.get("mac") or "").lower().strip()
        if dmac:
            _device_by_mac[dmac] = {
                "name": d.get("name") or d.get("hostname", ""),
                "type": _classify_device_type(d.get("type", "") or ""),
                "ip": d.get("ip", ""),
                "mac": dmac,
            }

    # Build reverse uplink lookup: (upstream_mac, port_idx) -> [device summaries]
    # Each managed device reports which port on its upstream device it is connected to.
    _uplinked_by_port: dict[tuple, list] = {}
    for d in raw_devices:
        uplink = d.get("uplink") or {}
        up_mac = (uplink.get("uplink_mac") or "").lower().strip()
        up_port = uplink.get("port_idx")
        if not up_mac or up_port is None:
            continue
        try:
            port_int = int(up_port)
        except (ValueError, TypeError):
            continue
        dmac = (d.get("mac") or "").lower().strip()
        _uplinked_by_port.setdefault((up_mac, port_int), []).append({
            "name": d.get("name") or d.get("hostname", ""),
            "type": _classify_device_type(d.get("type", "") or ""),
            "ip": d.get("ip", ""),
            "mac": dmac,
        })

    devices = []
    # port_tables: device_ip -> list of port_table entries from the UniFi API
    port_tables: dict[str, list] = {}
    for d in raw_devices:
        ip = d.get("ip", "")
        if not ip or d.get("state", 0) != 1:
            continue
        raw_type = d.get("type", "") or ""
        dtype = _classify_device_type(raw_type)
        devices.append({
            "ip": ip,
            "name": d.get("name") or d.get("hostname", ip),
            "type": dtype,
            "raw_type": raw_type,
            "mac": (d.get("mac") or "").lower().strip(),
        })
        if dtype == "switch":
            pts = d.get("port_table", [])
            if pts:
                port_tables[ip] = pts

    if not devices:
        raise HTTPException(status_code=404, detail="No connected devices found on controller")

    # 2. SSH into each device in parallel to fetch interfaces.
    #    SSH errors are caught internally; port_table data supplements/replaces missing ports.
    # Limit parallel SSH connections — small switches reject extra handshakes
    ssh_sem = asyncio.Semaphore(4)

    async def _fetch_for_device(device: dict) -> tuple[list, Optional[str]]:
        if device["type"] == "gateway":
            ssh_user = params.udm_ssh_username
            ssh_pass = params.udm_ssh_password
            ssh_key  = params.udm_ssh_key_path
        else:
            ssh_user = params.device_ssh_username
            ssh_pass = params.device_ssh_password
            ssh_key  = params.device_ssh_key_path
        capturer = PcapCapture(
            host=device["ip"],
            username=ssh_user,
            password=ssh_pass,
            key_path=ssh_key,
            port=params.ssh_port,
        )
        ssh_err: Optional[str] = None
        ifaces: list = []
        try:
            async with ssh_sem:
                ifaces = await asyncio.to_thread(capturer.fetch_interfaces, vlan_map or None)
        except Exception as e:
            ssh_err = str(e)

        is_switch = device["type"] == "switch"
        dev_mac = device["mac"]

        def _get_connected(port_idx_val) -> list:
            """Return clients connected to this device port (by device MAC + port index)."""
            if not dev_mac or port_idx_val is None:
                return []
            return _client_by_port.get((dev_mac, int(port_idx_val)), [])

        def _get_uplinked_devices(port_idx_val) -> list:
            """Return all UniFi devices uplinked to this port (switch/AP → this device)."""
            if not dev_mac or port_idx_val is None:
                return []
            try:
                pidx_int = int(port_idx_val)
            except (ValueError, TypeError):
                return []
            found: list = list(_uplinked_by_port.get((dev_mac, pidx_int), []))
            # Also check raw_clients for managed devices reporting sw_mac/sw_port
            for c in raw_clients:
                if (c.get("sw_mac") or "").lower() != dev_mac:
                    continue
                try:
                    if int(c.get("sw_port") or 0) != pidx_int:
                        continue
                except (ValueError, TypeError):
                    continue
                cmac = (c.get("mac") or "").lower().strip()
                if cmac in _device_by_mac:
                    dev_entry = _device_by_mac[cmac]
                    if not any(x["mac"] == cmac for x in found):
                        found.append(dev_entry)
            return found

        def _get_uplink_device(port_idx_val) -> Optional[dict]:
            """Return first uplinked UniFi device for backward compatibility."""
            devs = _get_uplinked_devices(port_idx_val)
            return devs[0] if devs else None

        for iface in ifaces:
            iface["device_ip"] = device["ip"]
            iface["device_name"] = device["name"]
            iface["device_type"] = device["type"]
            if is_switch:
                iface["capturable"] = False
            # Attach connected clients if port_idx known
            pidx = iface.get("port_idx")
            if pidx is not None:
                iface["connected_clients"] = _get_connected(pidx)
                updevs = _get_uplinked_devices(pidx)
                if updevs:
                    iface["connected_device"] = updevs[0]
                    iface["uplinked_devices"] = updevs

        # Supplement from UniFi API port_table (catches hw-switched ports SSH can't see)
        known_names = {i["name"] for i in ifaces}
        for port in port_tables.get(device["ip"], []):
            port_idx = port.get("port_idx")
            if not port_idx:
                continue
            ifname = port.get("ifname") or ""
            port_label = port.get("name") or f"Port {port_idx}"
            if ifname and ifname in known_names:
                # Enrich already-discovered interface with physical label
                for iface in ifaces:
                    if iface["name"] == ifname:
                        iface.setdefault("port_label", port_label)
                        iface.setdefault("port_idx", port_idx)
                        iface.setdefault("connected_clients", _get_connected(port_idx))
                        updevs = _get_uplinked_devices(port_idx)
                        if updevs:
                            iface.setdefault("connected_device", updevs[0])
                            iface.setdefault("uplinked_devices", updevs)
                continue
            # Use ifname if available, otherwise synthesize from port index
            # (hardware-switched ports may not have a Linux interface name)
            effective_name = ifname if ifname else f"port{port_idx}"
            if effective_name in known_names:
                continue
            speed = int(port.get("speed") or 0)
            _updev = _get_uplink_device(port_idx)
            _new_iface = {
                "name": effective_name,
                "type": "ethernet",
                "state": "up" if port.get("up") else "down",
                "mac": port.get("mac", ""),
                "mtu": 1500,
                "ips": [],
                "vlan_id": None,
                "parent": None,
                "master": None,
                "network_name": None,
                "bridge_vlans": [],
                "link_speed": speed if speed > 0 else None,
                "rx_bytes": int(port.get("rx_bytes") or 0),
                "tx_bytes": int(port.get("tx_bytes") or 0),
                "port_label": port_label,
                "port_idx": port_idx,
                "capturable": not is_switch,
                "hw_switched": not bool(ifname),
                "device_ip": device["ip"],
                "device_name": device["name"],
                "device_type": device["type"],
                "connected_clients": _get_connected(port_idx),
            }
            _updevs = _get_uplinked_devices(port_idx)
            if _updevs:
                _new_iface["connected_device"] = _updevs[0]
                _new_iface["uplinked_devices"] = _updevs
            ifaces.append(_new_iface)
            known_names.add(effective_name)

        return ifaces, ssh_err

    results = await asyncio.gather(*[_fetch_for_device(d) for d in devices], return_exceptions=True)

    combined = []
    errors = []
    for device, result in zip(devices, results):
        if isinstance(result, Exception):
            errors.append({"device_ip": device["ip"], "device_name": device["name"], "error": str(result)})
        else:
            ifaces, ssh_err = result
            combined.extend(ifaces)
            if ssh_err:
                errors.append({"device_ip": device["ip"], "device_name": device["name"],
                                "error": f"SSH failed (showing API port data): {ssh_err}", "partial": True})

    device_summary = [
        {"ip": d["ip"], "name": d["name"], "type": d["type"], "raw_type": d["raw_type"],
         "port_table_count": len(port_tables.get(d["ip"], []))}
        for d in devices
    ]
    return JSONResponse({"interfaces": combined, "vlan_map": vlan_map, "errors": errors,
                         "device_summary": device_summary})


@app.post("/api/devices")
async def list_devices(params: DeviceListParams):
    """Return APs, switches, and gateways from the UniFi controller."""
    client = UnifiClient(
        host=params.host,
        username=params.username,
        password=params.password,
        port=params.port,
        site=params.site,
    )
    try:
        client.login()
        devices = client.get_devices()
        client.logout()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to fetch devices: {e}")

    def _list_dev_type(raw_type: str) -> str:
        t = (raw_type or "").lower()
        if t.startswith("usw") or t.startswith("ubb"): return "switch"
        if t.startswith("uap") or t.startswith("uwa"): return "ap"
        if t.startswith("udm") or t.startswith("uxg") or t.startswith("usg") or t in ("ugw",): return "gateway"
        return t or "unknown"

    result = []
    for d in devices:
        raw_type = d.get("type", "") or ""
        result.append({
            "ip": d.get("ip", ""),
            "name": d.get("name") or d.get("hostname", ""),
            "type": _list_dev_type(raw_type),
            "model": d.get("model", ""),
            "mac": d.get("mac", ""),
            "connected": d.get("state", 0) == 1,
            "version": d.get("version", ""),
        })

    type_order = {"gateway": 0, "ap": 1, "switch": 2}
    result.sort(key=lambda d: (type_order.get(d["type"], 9), (d["name"] or "").lower()))
    return JSONResponse({"devices": result})


@app.get("/api/health")
async def health():
    return {"status": "ok"}


@app.get("/api/config")
async def get_config():
    """Return saved credentials (passwords decrypted in transit over localhost)."""
    return JSONResponse(load_config())


@app.post("/api/config")
async def post_config(cfg: SavedConfig):
    """Encrypt and persist credentials to ~/.unifi-analyzer/config.json."""
    save_config(cfg.model_dump())
    return {"status": "saved"}


# ---------------------------------------------------------------------------
# Live log streaming
# ---------------------------------------------------------------------------

@app.get("/api/logs/stream")
async def stream_logs(
    host: str = Query(...),
    ssh_username: str = Query(...),
    ssh_password: str = Query(""),
    ssh_port: int = Query(22),
    log_path: str = Query("/var/log/messages"),
):
    """
    Stream log lines via SSE (Server-Sent Events).
    Opens an SSH connection and runs `tail -f <log_path>`, pushing each line
    as an SSE `data:` event.  The browser closes the connection to stop streaming.
    """
    import paramiko

    line_queue: queue.Queue = queue.Queue(maxsize=2000)
    stop_event = threading.Event()

    def _ssh_worker():
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            client.connect(
                host,
                port=ssh_port,
                username=ssh_username,
                password=ssh_password or None,
                timeout=15,
                banner_timeout=30,
                auth_timeout=30,
                look_for_keys=False,
                allow_agent=False,
            )
            transport = client.get_transport()
            channel = transport.open_session()
            channel.settimeout(1.0)
            channel.exec_command(f"tail -f {log_path} 2>&1")

            buf = ""
            while not stop_event.is_set():
                try:
                    chunk = channel.recv(4096).decode("utf-8", errors="replace")
                    if not chunk:
                        break
                    buf += chunk
                    while "\n" in buf:
                        line, buf = buf.split("\n", 1)
                        line_queue.put(line.rstrip("\r"))
                except Exception:
                    pass
                if channel.exit_status_ready():
                    break
        except Exception as exc:
            line_queue.put(f"\x00ERROR\x00 {exc}")
        finally:
            try:
                client.close()
            except Exception:
                pass
            line_queue.put(None)  # sentinel

    thread = threading.Thread(target=_ssh_worker, daemon=True)
    thread.start()

    async def event_generator():
        loop = asyncio.get_event_loop()
        try:
            while True:
                try:
                    line = await loop.run_in_executor(None, lambda: line_queue.get(timeout=0.2))
                except queue.Empty:
                    yield ": keepalive\n\n"
                    continue
                if line is None:
                    break
                safe = line.replace("\n", " ").replace("\r", "")
                yield f"data: {safe}\n\n"
        except asyncio.CancelledError:
            pass
        finally:
            stop_event.set()

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


def _build_ports(devices_raw: list, portconf_raw: list) -> list:
    """Build port rows from UniFi stat/device and portconf data."""
    portconf_by_id = {
        p.get("_id"): p.get("name")
        for p in portconf_raw
        if isinstance(p, dict) and p.get("_id") and p.get("name")
    }
    rows = []
    for d in devices_raw:
        raw_type = (d.get("type") or "").lower()
        is_switch  = raw_type.startswith("usw") or raw_type.startswith("ubb")
        is_gateway = raw_type.startswith("udm") or raw_type.startswith("uxg") or raw_type in ("ugw", "usg")
        if not (is_switch or is_gateway):
            continue
        device_name = d.get("name") or d.get("hostname") or d.get("mac", "")
        port_overrides = {
            po.get("port_idx"): po
            for po in (d.get("port_overrides") or [])
            if isinstance(po, dict)
        }
        for port in (d.get("port_table") or []):
            if not isinstance(port, dict):
                continue
            port_idx = port.get("port_idx")
            override = port_overrides.get(port_idx, {})
            portconf_id = override.get("portconf_id") or port.get("portconf_id")
            row = {
                "_sort_gateway": 0 if is_gateway else 1,
                "_sort_device":  device_name.lower(),
                "device":  device_name,
                "port":    port_idx,
                "name":    port.get("name") or f"Port {port_idx}",
                "profile": portconf_by_id.get(portconf_id) if portconf_id else None,
                "enabled": port.get("enable"),
                "up":      port.get("up"),
                "speed":   port.get("speed"),
                "duplex":  "full" if port.get("full_duplex") is True else ("half" if port.get("full_duplex") is False else None),
                "media":   port.get("media"),
                "poe":     port.get("poe_mode") or override.get("poe_mode"),
            }
            rows.append({k: v for k, v in row.items() if v is not None})
    rows.sort(key=lambda r: (r.pop("_sort_gateway"), r.pop("_sort_device"), r.get("port") or 0))
    return rows


# ---------------------------------------------------------------------------
# Live ports fetch — switch + UDM port_table from stat/device
# ---------------------------------------------------------------------------

@app.get("/api/ports")
async def fetch_ports_live(
    host: str = Query(...),
    username: str = Query(...),
    password: str = Query(...),
    site: str = Query(default="default"),
):
    clean_host = re.sub(r'^https?://', '', host).rstrip('/')
    client = UnifiClient(host=clean_host, username=username, password=password, site=site)
    try:
        client.login()
        devices = client.get_devices()
        try:
            portconf_data = client._get("rest/portconf").get("data", [])
        except Exception:
            portconf_data = []
    except Exception as exc:
        raise HTTPException(status_code=502, detail=str(exc))
    finally:
        try:
            client.logout()
        except Exception:
            pass

    return {"ports": _build_ports(devices, portconf_data)}


# ---------------------------------------------------------------------------
# Config Export — run unifi_config_export.py as subprocess, stream output
# ---------------------------------------------------------------------------

@app.get("/api/export-config/run")
async def run_config_export(
    host: str = Query(...),
    username: str = Query(...),
    password: str = Query(...),
    humanize_epochs: bool = Query(False),
    strip_all_ids: bool = Query(False),
):
    """
    Run unifi config export in-process via SSE.
    Sends \x00JOBID\x00 first, then log lines, then \x00DONE\x00.
    """
    job_id = str(uuid.uuid4())
    out_dir = Path(tempfile.mkdtemp(prefix=f"unifi_export_{job_id[:8]}_"))
    url = f"https://{host}" if not host.startswith("http") else host

    line_queue: queue.Queue = queue.Queue(maxsize=5000)

    def _worker():
        try:
            run_export(
                url=url,
                user=username,
                password=password,
                out_dir=out_dir,
                log_fn=lambda msg: line_queue.put(str(msg)),
                humanize_epochs=humanize_epochs,
                strip_all_ids=strip_all_ids,
            )
            _export_jobs[job_id] = {"out_dir": out_dir, "exit_code": 0}
            line_queue.put(f"\x00DONE\x00 0 {job_id}")
        except Exception as exc:
            line_queue.put(f"\x00ERROR\x00 {traceback.format_exc()}")
            line_queue.put(f"\x00DONE\x00 1 {job_id}")
        finally:
            line_queue.put(None)

    threading.Thread(target=_worker, daemon=True).start()

    async def event_generator():
        loop = asyncio.get_event_loop()
        yield f"data: \x00JOBID\x00 {job_id}\n\n"
        try:
            while True:
                try:
                    line = await loop.run_in_executor(None, lambda: line_queue.get(timeout=0.2))
                except queue.Empty:
                    yield ": keepalive\n\n"
                    continue
                if line is None:
                    break
                safe = line.replace("\n", " ").replace("\r", "")
                yield f"data: {safe}\n\n"
        except asyncio.CancelledError:
            pass

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/config-lookup/fetch")
async def config_lookup_fetch(
    host: str = Query(...),
    username: str = Query(...),
    password: str = Query(...),
    site: str = Query(default="default"),
):
    """Fetch all Config Lookup data live from the UniFi controller using session-based auth."""
    def _join(val) -> str:
        if val is None:
            return ""
        if isinstance(val, list):
            return "; ".join(str(v) for v in val if v)
        return str(val)

    # --- Session-based client for standard endpoints ---
    try:
        client = UnifiClient(host=host, username=username, password=password, site=site)
        client.login()
    except Exception as exc:
        raise HTTPException(status_code=401, detail=f"Login failed: {exc}")

    try:
        networks_raw     = client.get_network_conf()
        wlans_raw        = client.get_wlan_conf()
        devices_raw      = client.get_devices()
        clients_raw      = client.get_clients()
        port_fw_raw      = client.get_port_forwards()
        routing_raw      = client.get_routing()
        try:
            port_profiles_raw = client._get("rest/portconf").get("data", [])
        except Exception:
            port_profiles_raw = []
        try:
            firewall_rules_raw = client.get_firewall_rules()
        except Exception:
            firewall_rules_raw = []
        try:
            firewall_zones_v2_raw = client.get_firewall_zones_v2_session()
        except Exception:
            firewall_zones_v2_raw = []
        try:
            firewall_policies_raw = client.get_firewall_policies_session()
        except Exception:
            firewall_policies_raw = []
        try:
            dns_records_raw = client.get_dns_records_session()
        except Exception:
            dns_records_raw = []
        try:
            firewall_groups_raw = client.get_firewall_groups()
        except Exception:
            firewall_groups_raw = []
        client.logout()
    except Exception as exc:
        try:
            client.logout()
        except Exception:
            pass
        raise HTTPException(status_code=502, detail=f"Failed to fetch data: {exc}")

    # --- Old-style firewall rules (classic firewall, pre-zone-based) ---
    firewall_rules_session: list = [
        {
            "name":        r.get("name", ""),
            "enabled":     r.get("enabled"),
            "action":      r.get("action", ""),
            "protocol":    r.get("protocol") or "all",
            "logging":     r.get("log") or r.get("logging"),
            "src_zone":    r.get("ruleset", ""),
            "dst_zone":    "",
            "src_address": _join(r.get("src_ip") or r.get("src_cidr")),
            "dst_address": _join(r.get("dst_ip") or r.get("dst_cidr")),
            "src_port":    _join(r.get("src_port")),
            "dst_port":    _join(r.get("dst_port")),
            "description": r.get("comment") or r.get("description", ""),
        }
        for r in firewall_rules_raw if isinstance(r, dict)
    ]

    # Build network ID → name lookup from already-fetched session data
    network_id_to_name: dict = {
        n["_id"]: n.get("name", n["_id"])
        for n in networks_raw if isinstance(n, dict) and n.get("_id")
    }

    # Build firewall zones from v2 session API (zone_id = _id, networks = network_ids)
    firewall_zones: list = [
        {
            "zone_id":  z.get("_id") or z.get("id", ""),
            "name":     z.get("name", ""),
            "networks": "; ".join(
                network_id_to_name.get(mid, mid)
                for mid in (z.get("network_ids") or z.get("members") or [])
                if mid
            ),
        }
        for z in firewall_zones_v2_raw if isinstance(z, dict)
    ]

    firewall_policies: list = []
    integration_error: str = ""

    _excl_pc = {"native_networkconf_id", "port_security_mac_address_site_id", "_id", "site_id"}
    _excl_rt = {"site_id", "_id"}

    # Zone ID → name lookup for resolving policy source/destination zone IDs
    _zone_id_to_name: dict = {z["zone_id"]: z["name"] for z in firewall_zones if z.get("zone_id")}

    return {
        "site": site,
        "integration_error": integration_error,
        "networks": [
            {"name": n.get("name"), "purpose": n.get("purpose"),
             "ip_subnet": n.get("ip_subnet"), "enabled": n.get("enabled")}
            for n in networks_raw if isinstance(n, dict) and n.get("name")
        ],
        "wifi": [
            {"name": n.get("name"), "security": n.get("security"), "band": n.get("wlan_band"),
             "vlan_id": n.get("vlanid"), "enabled": n.get("enabled"), "hide_ssid": n.get("hide_ssid")}
            for n in wlans_raw if isinstance(n, dict)
        ],
        "devices": [
            {"name": d.get("name") or d.get("hostname"), "model": d.get("model_name") or d.get("model"),
             "firmware": d.get("version"), "mac": d.get("mac"), "ip": d.get("ip"), "type": d.get("type"),
             "state": d.get("state"), "uptime": d.get("uptime")}
            for d in devices_raw if isinstance(d, dict)
        ],
        "clients": [
            {"hostname": c.get("hostname") or c.get("name"), "ip": c.get("ip"),
             "mac": c.get("mac"), "network": c.get("network"), "manufacturer": c.get("oui"),
             "is_wired": c.get("is_wired"), "uptime": c.get("uptime")}
            for c in clients_raw if isinstance(c, dict)
        ],
        "firewall": firewall_policies or [
            {
                "name":        p.get("name", ""),
                "enabled":     p.get("enabled"),
                "action":      (p.get("action", "") or "").upper(),
                "protocol":    p.get("protocol") or "all",
                "logging":     p.get("logging"),
                "src_zone":    _zone_id_to_name.get(
                                   (p.get("source") or {}).get("zone_id", ""),
                                   (p.get("source") or {}).get("zone_id", "")),
                "dst_zone":    _zone_id_to_name.get(
                                   (p.get("destination") or {}).get("zone_id", ""),
                                   (p.get("destination") or {}).get("zone_id", "")),
                "src_address": _join((p.get("source") or {}).get("ips") or
                                     (p.get("source") or {}).get("matching_target")),
                "dst_address": _join((p.get("destination") or {}).get("ips") or
                                     (p.get("destination") or {}).get("matching_target")),
                "src_port":    str((p.get("source") or {}).get("port", "") or ""),
                "dst_port":    str((p.get("destination") or {}).get("port", "") or ""),
                "description": p.get("description", ""),
            }
            for p in firewall_policies_raw if isinstance(p, dict)
        ] or firewall_rules_session,
        "firewall_zones": firewall_zones,
        "dns_records": [
            {
                "domain":      r.get("key", ""),
                "type":        r.get("record_type", "A"),
                "value":       r.get("value", ""),
                "ttl":         "Auto" if r.get("ttl", 0) == 0 else str(r.get("ttl")) + "s",
                "enabled":     r.get("enabled"),
                "priority":    r.get("priority"),
                "weight":      r.get("weight"),
                "port":        r.get("port"),
            }
            for r in dns_records_raw if isinstance(r, dict)
        ],
        "port_forward": [
            {
                "name":           p.get("name", ""),
                "enabled":        p.get("enabled"),
                "proto":          p.get("proto") or p.get("protocol", "tcp_udp"),
                "dst_port":       p.get("dst_port") or p.get("fwd_outside_port", ""),
                "fwd":            p.get("fwd") or p.get("fwd_ip", ""),
                "fwd_port":       p.get("fwd_port", ""),
                "src":            p.get("src", ""),
                "interface":      p.get("pfwd_interface", ""),
                "destination_ip": "" if p.get("destination_ip", "") in ("", "any") else p.get("destination_ip"),
                "logging":        p.get("log"),
            }
            for p in port_fw_raw if isinstance(p, dict)
        ],
        "policy_groups": [
            {
                "name":    g.get("name", ""),
                "type":    g.get("group_type", ""),
                "members": "; ".join(str(m) for m in g.get("group_members", []) if m),
                "count":   len(g.get("group_members", [])),
            }
            for g in firewall_groups_raw if isinstance(g, dict)
        ],
        "port_profiles": [
            {k: v for k, v in p.items() if k not in _excl_pc}
            for p in port_profiles_raw if isinstance(p, dict)
        ],
        "routing": [
            {k: v for k, v in r.items() if k not in _excl_rt}
            for r in routing_raw if isinstance(r, dict)
        ],
        "ports": _build_ports(devices_raw, port_profiles_raw),
    }


@app.get("/api/export-config/preview/{job_id}")
async def preview_export(job_id: str):
    """Return structured preview data from a completed export job."""
    job = _export_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Export job not found or not yet complete")
    out_dir: Path = job["out_dir"]

    def _unifi_data(path: Path) -> list:
        """Read a saved UniFi API JSON file and return the inner data list."""
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
            # Files are saved as {"status_code":..., "payload": <api response>}
            # The API response is typically {"data": [...], "meta": {...}}
            payload = raw.get("payload", raw) if isinstance(raw, dict) else raw
            if isinstance(payload, dict):
                data = payload.get("data", [])
            elif isinstance(payload, list):
                data = payload
            else:
                data = []
            return data if isinstance(data, list) else []
        except Exception:
            return []

    def _load(filename: str) -> list:
        p = site_dir / filename
        return _unifi_data(p) if p.exists() else []

    def _join(values) -> str:
        if not values:
            return ""
        if isinstance(values, list):
            return "; ".join(str(v) for v in values if v)
        return str(values)

    def _build_switch_ports(site_dir: Path, devices_raw: list) -> list:
        """Extract per-port interface data from integration API device responses."""
        int_devices = _unifi_data(site_dir / "integration_devices.json") if (site_dir / "integration_devices.json").exists() else []

        rows = []
        for dev in int_devices:
            if not isinstance(dev, dict):
                continue
            interfaces = dev.get("interfaces") or []
            if not interfaces:
                continue
            device_name = dev.get("name") or dev.get("hostname") or dev.get("id", "")
            for iface in interfaces:
                if not isinstance(iface, dict):
                    continue
                row = {"device_name": device_name}
                row.update(iface)
                rows.append(row)
        return rows

    def _build_firewall_rules(site_dir: Path) -> list:
        # Build group lookup from rest_firewallgroup.json
        groups_raw = _unifi_data(site_dir / "rest_firewallgroup.json") if (site_dir / "rest_firewallgroup.json").exists() else []
        group_by_id: dict = {}
        for g in groups_raw:
            if not isinstance(g, dict):
                continue
            gid = g.get("_id")
            if gid:
                group_by_id[gid] = {
                    "name": g.get("name", ""),
                    "members": g.get("group_members") or g.get("members") or [],
                }

        def resolve_groups(ids) -> tuple:
            if not ids:
                return "", ""
            if isinstance(ids, str):
                ids = [ids]
            names, members = [], []
            for gid in ids:
                g = group_by_id.get(gid)
                if g:
                    if g["name"]:
                        names.append(g["name"])
                    members.extend(str(m) for m in g["members"] if m)
            return "; ".join(names), "; ".join(members)

        rules_raw = _unifi_data(site_dir / "rest_firewallrule.json") if (site_dir / "rest_firewallrule.json").exists() else []
        rows = []
        for r in rules_raw:
            if not isinstance(r, dict):
                continue
            src_group_ids = r.get("src_firewall_group_ids") or r.get("src_group_ids") or []
            dst_group_ids = r.get("dst_firewall_group_ids") or r.get("dst_group_ids") or []
            src_group_names, src_group_members = resolve_groups(src_group_ids)
            dst_group_names, dst_group_members = resolve_groups(dst_group_ids)
            rows.append({
                "name":         r.get("name") or "",
                "ruleset":      r.get("ruleset") or "",
                "index":        r.get("rule_index") or r.get("index") or "",
                "action":       r.get("action") or "",
                "protocol":     r.get("protocol") or "all",
                "enabled":      r.get("enabled"),
                "logging":      r.get("log") or r.get("logging"),
                "src_address":  _join(r.get("src_ip") or r.get("src_cidr")),
                "src_port":     _join(r.get("src_port")),
                "src_groups":   src_group_names,
                "src_members":  src_group_members,
                "dst_address":  _join(r.get("dst_ip") or r.get("dst_cidr")),
                "dst_port":     _join(r.get("dst_port")),
                "dst_groups":   dst_group_names,
                "dst_members":  dst_group_members,
                "description":  r.get("comment") or r.get("description") or "",
            })
        return rows

    sections = {}

    for site_dir in sorted(out_dir.iterdir()):
        if not site_dir.is_dir():
            continue
        site_name = site_dir.name

        summary = {}
        summary_file = site_dir / "_summary.json"
        if summary_file.exists():
            try:
                summary = json.loads(summary_file.read_text(encoding="utf-8"))
            except Exception:
                pass

        networks_raw = _load("rest_networkconf.json")
        networks = [{"name": n.get("name"), "purpose": n.get("purpose"),
                     "ip_subnet": n.get("ip_subnet"), "enabled": n.get("enabled")}
                    for n in networks_raw if isinstance(n, dict) and n.get("name")]

        wlans_raw = _load("rest_wlanconf.json")
        wifi = [{"name": n.get("name"), "security": n.get("security"), "band": n.get("wlan_band"),
                 "vlan_id": n.get("vlanid"), "enabled": n.get("enabled"),
                 "hide_ssid": n.get("hide_ssid")} for n in wlans_raw if isinstance(n, dict)]

        devices_raw = _load("stat_device.json")
        devices = [{"name": d.get("name") or d.get("hostname"), "model": d.get("model_name") or d.get("model"),
                    "firmware": d.get("version"), "mac": d.get("mac"), "ip": d.get("ip"), "type": d.get("type"),
                    "state": d.get("state"), "uptime": d.get("uptime")} for d in devices_raw if isinstance(d, dict)]

        clients_raw = _load("stat_sta.json")
        clients = [{"hostname": c.get("hostname") or c.get("name"), "ip": c.get("ip"),
                    "mac": c.get("mac"), "network": c.get("network"), "manufacturer": c.get("oui"),
                    "is_wired": c.get("is_wired"), "uptime": c.get("uptime")} for c in clients_raw if isinstance(c, dict)]

        firewall = _build_firewall_rules(site_dir)
        switch_ports = _build_switch_ports(site_dir, devices_raw)

        port_fw_raw = _load("rest_portforward.json") or _load("stat_portforward.json")
        port_fw = [{"name": p.get("name"), "proto": p.get("proto"),
                    "dst_port": p.get("dst_port") or p.get("fwd_outside_port"),
                    "fwd": p.get("fwd"), "fwd_port": p.get("fwd_port"),
                    "enabled": p.get("enabled")} for p in port_fw_raw if isinstance(p, dict)]

        _excl_pc = {"native_networkconf_id", "port_security_mac_address_site_id", "_id"}
        port_profiles = [{k: v for k, v in p.items() if k not in _excl_pc}
                         for p in _load("rest_portconf.json") if isinstance(p, dict)]

        _excl_rt = {"site_id", "_id"}
        routing = [{k: v for k, v in r.items() if k not in _excl_rt}
                   for r in _load("rest_routing.json") if isinstance(r, dict)]

        sections[site_name] = {
            "summary": summary,
            "networks": networks,
            "wifi": wifi,
            "devices": devices,
            "clients": clients,
            "firewall": firewall,
            "port_forward": port_fw,
            "port_profiles": port_profiles,
            "ports": switch_ports,
            "routing": routing,
        }

    return {"sites": sections}


@app.get("/api/export-config/download/{job_id}/excel")
async def download_export_excel(job_id: str):
    """Stream the generated Excel file(s) from a completed export job."""
    job = _export_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Export job not found or not yet complete")
    out_dir: Path = job["out_dir"]
    xlsx_files = list(out_dir.rglob("*.xlsx"))
    if not xlsx_files:
        raise HTTPException(status_code=404, detail="No Excel file found — ensure pandas/openpyxl are installed")

    if len(xlsx_files) == 1:
        data = xlsx_files[0].read_bytes()
        async def _stream_xlsx():
            yield data
        return StreamingResponse(
            _stream_xlsx(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{xlsx_files[0].name}"'},
        )

    # Multiple xlsx — bundle into a zip
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for f in xlsx_files:
            zf.writestr(f.name, f.read_bytes())
    zip_bytes = buf.getvalue()
    async def _stream_zip():
        yield zip_bytes
    return StreamingResponse(
        _stream_zip(),
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="unifi_export_excel.zip"'},
    )


@app.get("/api/export-config/download/{job_id}/json")
async def download_export_json(job_id: str):
    """Zip all JSON files from a completed export job and stream to browser."""
    job = _export_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Export job not found or not yet complete")
    out_dir: Path = job["out_dir"]
    json_files = list(out_dir.rglob("*.json"))
    if not json_files:
        raise HTTPException(status_code=404, detail="No JSON files found in export output")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for f in json_files:
            arcname = str(f.relative_to(out_dir))
            zf.writestr(arcname, f.read_bytes())
    zip_bytes = buf.getvalue()

    async def _stream():
        yield zip_bytes
    return StreamingResponse(
        _stream(),
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="unifi_export_json.zip"'},
    )


# ---------------------------------------------------------------------------
# History endpoints
# ---------------------------------------------------------------------------

from fastapi import UploadFile, File, Form as FormField

@app.post("/api/history/save")
async def history_save(
    file: UploadFile = File(...),
    udm_name: str = FormField(default="udm"),
    site: str = FormField(default="default"),
):
    """Save a generated Excel file to the history directory."""
    HISTORY_DIR.mkdir(exist_ok=True)
    contents = await file.read()
    # Sanitise and build filename
    safe_name = re.sub(r"[^a-zA-Z0-9._-]", "_", udm_name) or "udm"
    from datetime import datetime
    ts = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    filename = f"{safe_name}-{ts}.xlsx"
    filepath = HISTORY_DIR / filename
    filepath.write_bytes(contents)
    stat = filepath.stat()
    return {
        "filename": filename,
        "saved_at": ts,
        "size_bytes": stat.st_size,
        "udm_name": udm_name,
        "site": site,
    }


@app.post("/api/history/save-pcap")
async def history_save_pcap(
    file: UploadFile = File(...),
    udm_name: str = FormField(default="udm"),
    interface: str = FormField(default="eth0"),
):
    """Save a raw PCAP file to the history directory."""
    HISTORY_DIR.mkdir(exist_ok=True)
    contents = await file.read()
    safe_name  = re.sub(r"[^a-zA-Z0-9._-]", "_", udm_name.strip())  or "udm"
    safe_iface = re.sub(r"[^a-zA-Z0-9._-]", "_", interface.strip()) or "eth0"
    from datetime import datetime
    ts = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    filename = f"{safe_name}-{safe_iface}-{ts}.pcap"
    filepath = HISTORY_DIR / filename
    filepath.write_bytes(contents)
    return {
        "filename": filename,
        "saved_at": ts,
        "size_bytes": filepath.stat().st_size,
        "udm_name": safe_name,
        "interface": safe_iface,
    }


@app.get("/api/history")
async def history_list():
    """Return metadata for all saved history files (xlsx and pcap)."""
    from datetime import datetime
    def _meta(f):
        stat = f.stat()
        return {"filename": f.name,
                "saved_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "size_bytes": stat.st_size}
    if not HISTORY_DIR.exists():
        return {"files": [], "pcap_files": []}
    files = [_meta(f) for f in sorted(HISTORY_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True)]
    pcap_files = [_meta(f) for f in sorted(HISTORY_DIR.glob("*.pcap"), key=lambda x: x.stat().st_mtime, reverse=True)]
    return {"files": files, "pcap_files": pcap_files}


@app.get("/api/history/download/{filename}")
async def history_download(filename: str):
    """Download a history file by name."""
    # Prevent path traversal
    safe = re.sub(r"[^a-zA-Z0-9._-]", "_", filename)
    filepath = HISTORY_DIR / safe
    if not filepath.exists() or filepath.suffix not in (".xlsx", ".pcap"):
        raise HTTPException(status_code=404, detail="File not found")
    data = filepath.read_bytes()
    media = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
             if filepath.suffix == ".xlsx" else "application/vnd.tcpdump.pcap")
    async def _stream():
        yield data
    return StreamingResponse(
        _stream(),
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{safe}"'},
    )


@app.delete("/api/history")
async def history_delete(body: dict):
    """Delete one or more history files by filename."""
    filenames = body.get("filenames", [])
    deleted, errors = [], []
    for name in filenames:
        safe = re.sub(r"[^a-zA-Z0-9._-]", "_", name)
        filepath = HISTORY_DIR / safe
        if filepath.exists() and filepath.suffix in (".xlsx", ".pcap"):
            filepath.unlink()
            deleted.append(safe)
        else:
            errors.append(safe)
    return {"deleted": deleted, "errors": errors}


@app.post("/api/history/download-zip")
async def history_download_zip(body: dict):
    """Return a zip archive containing one or more history files."""
    filenames = body.get("filenames", [])
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in filenames:
            safe = re.sub(r"[^a-zA-Z0-9._-]", "_", name)
            filepath = HISTORY_DIR / safe
            if filepath.exists() and filepath.suffix in (".xlsx", ".pcap"):
                zf.write(filepath, safe)
    buf.seek(0)

    async def _stream():
        yield buf.read()

    return StreamingResponse(
        _stream(),
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="history-export.zip"'},
    )


@app.post("/api/history/compare")
async def history_compare(body: dict):
    """
    Compare selected history files against the most recent history file.

    Summary sheet (pivoted): sections as rows, each compared file as a column.
    One diff sheet per section that has any differences — columns are the data
    fields plus one "vs {file}" status column per compared file.
    Green = Added in latest (not in that older file), Red = Removed from latest,
    Yellow = mixed across compared files.
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from datetime import datetime

    filenames = body.get("filenames", [])
    if not HISTORY_DIR.exists():
        raise HTTPException(status_code=404, detail="No history files found")

    all_files = sorted(HISTORY_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True)
    if not all_files:
        raise HTTPException(status_code=404, detail="No history files found")

    latest_file   = all_files[0]
    latest_xl     = pd.ExcelFile(latest_file)
    latest_sheets = {s: latest_xl.parse(s) for s in latest_xl.sheet_names}

    GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")   # Added (in latest, not compared)
    RED_FILL    = PatternFill("solid", fgColor="FFC7CE")   # Removed (in compared, not latest)
    YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")   # Mixed across compared files
    HEADER_FILL = PatternFill("solid", fgColor="F97316")   # Orange header
    WHITE_BOLD  = Font(bold=True, color="FFFFFF")

    def make_short_name(fname: str) -> str:
        base = fname.replace(".xlsx", "").replace("_", " ")
        if len(base) <= 30:
            return base
        return base[:10] + "\u2026" + base[-19:]

    # Build validated list of compared files (skip latest, skip invalid)
    compared_files: list = []
    short_names:    list = []
    for fname in filenames:
        safe = re.sub(r"[^a-zA-Z0-9._-]", "_", fname)
        filepath = HISTORY_DIR / safe
        if not filepath.exists() or filepath.suffix != ".xlsx":
            continue
        if filepath.resolve() == latest_file.resolve():
            continue
        compared_files.append(filepath)
        short_names.append(make_short_name(safe))

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    # Early exit when nothing to compare
    if not compared_files:
        summary_ws.append(["No comparisons possible — all selected files are the latest version"])
        for cell in summary_ws[1]:
            cell.fill = HEADER_FILL
            cell.font = WHITE_BOLD
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ts = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        async def _stream_empty():
            yield buf.read()
        return StreamingResponse(
            _stream_empty(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="comparison-{ts}.xlsx"'},
        )

    # Parse all compared workbooks once before the section loop
    compared_xl_list = []
    for fp in compared_files:
        cxl = pd.ExcelFile(fp)
        compared_xl_list.append({s: cxl.parse(s) for s in cxl.sheet_names})

    # Summary sheet — pivoted: rows = sections, columns = compared files
    summary_ws.append(["Section"] + short_names)
    for cell in summary_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD

    n_files = len(compared_files)

    for section_name in latest_xl.sheet_names:
        latest_df = latest_sheets[section_name].astype(str)
        data_cols = list(latest_df.columns)

        file_added:   list = []
        file_removed: list = []

        for comp_sheets in compared_xl_list:
            comp_df = comp_sheets.get(section_name, pd.DataFrame())
            if comp_df.empty:
                file_added.append(pd.DataFrame(columns=data_cols))
                file_removed.append(pd.DataFrame(columns=data_cols))
                continue
            try:
                # Normalise compared schema to match latest (handles added/removed columns)
                comp_aligned = comp_df.astype(str).reindex(columns=data_cols, fill_value="")
                merged  = pd.merge(latest_df, comp_aligned, how="outer", indicator=True)
                added   = merged[merged["_merge"] == "left_only"].drop("_merge", axis=1)
                removed = merged[merged["_merge"] == "right_only"].drop("_merge", axis=1)
            except Exception:
                added   = pd.DataFrame(columns=data_cols)
                removed = pd.DataFrame(columns=data_cols)
            file_added.append(added)
            file_removed.append(removed)

        # Build summary row for this section
        summary_cells = [section_name]
        for i in range(n_files):
            n_a, n_r = len(file_added[i]), len(file_removed[i])
            if not n_a and not n_r:
                summary_cells.append("No changes")
            else:
                parts = []
                if n_a: parts.append(f"{n_a} added")
                if n_r: parts.append(f"{n_r} removed")
                summary_cells.append(", ".join(parts))
        summary_ws.append(summary_cells)

        # Skip diff sheet if no differences exist for this section
        if not any(len(file_added[i]) or len(file_removed[i]) for i in range(n_files)):
            continue

        # Collect deduplicated differing rows across all compared files
        # diff_rows: {row_tuple: {file_idx: "Added" | "Removed"}}
        diff_rows: dict = {}
        for i in range(n_files):
            for _, row in file_added[i].iterrows():
                key = tuple(str(row.get(c, "")) for c in data_cols)
                diff_rows.setdefault(key, {})[i] = "Added"
            for _, row in file_removed[i].iterrows():
                key = tuple(str(row.get(c, "")) for c in data_cols)
                diff_rows.setdefault(key, {})[i] = "Removed"

        # Create one diff sheet per section (names are all ≤17 chars — no truncation needed)
        ws = wb.create_sheet(section_name[:31])
        vs_headers = [f"vs {short_names[i]}" for i in range(n_files)]
        ws.append(data_cols + vs_headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = WHITE_BOLD

        for row_tuple, statuses in diff_rows.items():
            vals = set(statuses.values())
            if "Added" in vals and "Removed" in vals:
                row_fill = YELLOW_FILL
            elif "Added" in vals:
                row_fill = GREEN_FILL
            else:
                row_fill = RED_FILL

            status_cells = [statuses.get(i, "") for i in range(n_files)]
            ws.append(list(row_tuple) + status_cells)
            for cell in ws[ws.max_row]:
                cell.fill = row_fill

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    # Auto-fit summary sheet columns
    for col in summary_ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        summary_ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    report_name = f"comparison-{ts}.xlsx"

    async def _stream():
        yield buf.read()

    return StreamingResponse(
        _stream(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{report_name}"'},
    )
